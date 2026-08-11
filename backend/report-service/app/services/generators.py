"""Generates real PDF (reportlab) and Excel (openpyxl) scientific reports from
project/experiment/prediction data — not placeholder files."""
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def build_pdf_report(title: str, sections: list[dict]) -> bytes:
    """sections: [{"heading": str, "rows": [[col1, col2, ...], ...]}]"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    story.append(Paragraph(f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}", styles["Normal"]))
    story.append(Spacer(1, 20))

    for section in sections:
        story.append(Paragraph(section["heading"], styles["Heading2"]))
        story.append(Spacer(1, 8))
        if section.get("rows"):
            table = Table(section["rows"])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), "#0B3D91"),
                        ("TEXTCOLOR", (0, 0), (-1, 0), "#FFFFFF"),
                        ("GRID", (0, 0), (-1, -1), 0.5, "#CCCCCC"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ]
                )
            )
            story.append(table)
        story.append(Spacer(1, 16))

    doc.build(story)
    return buffer.getvalue()


def build_excel_report(title: str, sheets: dict[str, list[list]]) -> bytes:
    """sheets: {"Sheet name": [[header...], [row...], ...]}"""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        for row in rows:
            ws.append(row)
        if rows:
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
